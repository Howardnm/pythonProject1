#!/usr/bin/env python
# -*- coding: utf-8 -*-
import music_tag
import os
import time
import re
import csv
import codecs
import pandas as pd
import openpyxl
import xlsxwriter
# import shutil
import colorama
from openpyxl.styles import Font  # 导入字体模块
from openpyxl.styles import PatternFill  # 导入填充模块

colorama.init(autoreset=True)  # exe打包文件print出彩色
################################################################
# 可以转任意编码到utf-8编码
import chardet
from chardet.universaldetector import UniversalDetector

################################################################

# 去cmd构建exe："pyinstaller -i 456.ico -F C:\Users\85099\PycharmProjects\pythonProject1\notion\专辑规范化审核工具.py"
# pip3 install requests
print("----------------------------------")
print(
    "wiki指南：\033[1;34mhttps://antique-hacksaw-c2a.notion.site/Plex-music-family-029133efa133486db0c3624e4c5521de\033[0m")

print("请确保音乐文件通过MusicTag规范化，不懂操作请看wiki指南：\033[1;35m【必读】音乐文件规范化\033[0m")
print("---------------------------------")
print("\033[31m欢迎使用plex专辑规范化自动审核 v1.0.6\033[0m")
print("\033[33m新增功能：lrc歌词文件自动修复乱码\033[0m")
print("---------------------------------")
print("请输入专辑文件夹绝对路径（在本窗口里右键，即可粘贴内容）")
print("\033[1;35m例如：C:" + "\\Users\85099\Desktop\G.E.M.邓紫棋 - 启示录\033[0m")
print("\033[1;35m例如：\\" + "\\192.168.1.20\downloads\音乐\G.E.M.邓紫棋 - 启示录\033[0m")
print("---------------------------------")
print("\033[33m把专辑文件夹拖进来\033[31m（若路径有双引号，请去掉）\033[0m")
music_files=""
def music_files_input():
    global music_files
    music_files=input("输入后回车：")
    if "“" in music_files or "”" in music_files or '"' in music_files:
        print("\033[31m输入有误，重新输入\033[0m")
        music_files_input()
music_files_input()


token='secret_DkYKff6IQJdGJTqgFslKQqGpLtsu44RiH3Vsa2nhYjO'

databaseId='7aaf52b505fd4b86ac164ad52d0f8570'

headers={
    "Authorization": "Bearer " + token,
    "Content-Type": "application/json",
    "Notion-Version": "2022-06-28",  # 2021-05-13
    'Cache-Control': 'no-cache'
}

music_list_cache1=music_files + "\\foo3.txt"
# print(music_list_cache1)
music_list_cache2=music_files + "\\foo4.txt"

today=time.strftime("%Y/%m/%d %H:%M", time.localtime())
print(today)
csv_title=["ID", "上层路径", "歌手", "歌曲", "音轨号", "专辑", "专辑艺术家", "年份", "歌词", "标签", "文件夹结构", "合 唱", "文件路径", today]
csv_file=music_files + "\\auto_check.csv"
xlsx_file=music_files + "\\自动审核表.xlsx"

# ----------------------------------------------------------------


music_list=[]  # 读取绝对路径，识别"/"，将列表内元素细分，形成套娃列表
music_list_AP=[]  # 读取绝对路径，写入列表
csv_dict=[]
num=0
GB_size=""
json_file1=""
music_tag_format1, music_tag_album, music_tag_year="", "", ""


########################################################################
# 获取文件编码类型，进行任意文件utf8编码
def get_encoding(file):
    # 二进制方式读取，获取字节数据，检测类型
    with open(file, 'rb') as f:
        data=f.read()
        return chardet.detect(data)['encoding']


def get_encode_info(file):
    with open(file, 'rb') as f:
        detector=UniversalDetector()
        for line in f.readlines():
            detector.feed(line)
            if detector.done:
                break
        detector.close()
        return detector.result['encoding']


def read_file(file):
    with open(file, 'rb') as f:
        return f.read()


def write_file(content, file):
    with open(file, 'wb') as f:
        f.write(content)


def convert_encode2utf8(file, original_encode, des_encode):
    file_content=read_file(file)
    file_decode=file_content.decode(original_encode, 'ignore')
    file_encode=file_decode.encode(des_encode)
    write_file(file_encode, file)


def encodeFile2Utf8(filename):
    file_content=read_file(filename)
    encode_info=get_encode_info(filename)
    if encode_info != 'utf-8':
        convert_encode2utf8(filename, encode_info, 'utf-8')


############################################################################
# 创建做种文件
def mkdir(path):
    folder=os.path.exists(path)

    if not folder:  # 判断是否存在文件夹如果不存在则创建为文件夹
        os.makedirs(path)  # makedirs 创建文件时如果路径不存在会创建这个路径
        print("---  创建做种文件夹  ---")

    else:
        print("---  做种文件夹已存在!  ---")


try:
    print(" ")
    # shutil.rmtree(music_files+"(适合做种的文件夹结构)") # 删除非空文件夹
except Exception as err:
    print(err)
    print("没有做种文件夹")


# mkdir(music_files+"(适合做种的文件夹结构)")
#############################################################################


# 遍历本地音乐文件
def find_music():
    print("\033[36m运行find_music\033[0m")
    list1=[]

    def gci(filepath):
        ppp=''
        # 遍历filepath下所有文件，包括子目录
        files=os.listdir(filepath)
        for fi in files:
            fi_d=os.path.join(filepath, fi)
            if os.path.isdir(fi_d):
                gci(fi_d)
            else:
                # print(os.path.join(filepath, fi_d))
                ppp=os.path.join(filepath, fi_d)
                list1.append(ppp)
                # print(list1)

    # 递归遍历/root目录下所有文件
    gci(music_files)

    # 遍历后的绝对路径，写入txt
    fo=open(music_list_cache1, "w+", encoding='utf8')
    for line in list1:
        fo.writelines(line + "\n")
    fo.close()
    # print(list1)

    # 删除包含".sync .txt .jpg .csv"的行
    re_str=['.sync', '.txt', '.jpg', '.csv', '.xlsx', '@eaDir']
    for i in range(len(re_str)):
        list2=[]
        matchPattern=re.compile(r'{}'.format(re_str[i]))
        file=open(music_list_cache1, 'r', encoding='utf-8')
        while 1:
            line=file.readline()
            if not line:
                print("删除列表中包含" + str(re_str[i]) + "的行")
                break
            elif matchPattern.search(line):
                pass
            else:
                list2.append(line)
        file.close()
        file=open(music_list_cache1, 'w', encoding='utf-8')
        for i in list2:
            file.write(i)
        file.close()

    file=open(music_list_cache1, 'w', encoding='utf-8')
    for i in list2:
        file.write(i)
    file.close()

    # 读取绝对路径，写入列表
    fo=open(music_list_cache1, "r", encoding='utf-8')
    for line in fo:
        line=line.replace("\n", "")
        music_list_AP.append(line)
    fo.close()
    print("----------------")
    # print(music_list_AP)
    # 读取绝对路径，识别"/"，将列表内元素细分，形成套娃列表
    fo=open(music_list_cache1, "r", encoding='utf-8')
    for line in fo:
        line=line.replace("\n", "")
        music_list.append(line.split("\\"))
    print(music_list)
    fo.close()
    # 整理后的列表重新写入
    fo=open(music_list_cache2, "w+", encoding='utf8')
    fo.write(str(music_list))
    fo.close()


# notion

num_true=0


def createPage(databaseId, headers):
    print("\033[1;35m运行createPage写表\033[0m")
    global num_true, num1, lyric1, Name_format, music_tag_format1, music_tag_album, music_tag_year, music_tag_chorus
    print("运行次数：" + str(num1))
    num_true+=1
    # print(num_true)
    music_file_path=f"/{music_list[num1][-4]}/{music_list[num1][-3]}/{music_list[num1][-2]}/{music_list[num1][-1]}"
    # music_CRC32=str(zlib.crc32(music_file_path.encode('utf8')))
    # print("music_CRC32=" + music_CRC32)
    music_CRC32=num_true  # ID换回序号

    to_csv=1
    if to_csv == 1:
        print("\033[1;35m歌曲信息导入字典数：" + str(num_true) + "\033[0m")
        global csv_dict
        global csv_title
        csv_dict_i={csv_title[0]: music_CRC32,
                    csv_title[1]: music_list[num1][music_area],
                    csv_title[2]: str(f['artist']),
                    csv_title[3]: str(f['title']),
                    csv_title[4]: music_tag_tracknumber,
                    csv_title[5]: music_tag_album,
                    csv_title[6]: music_tag_albumartist,
                    csv_title[7]: music_tag_year,
                    csv_title[8]: str(lyric1),
                    csv_title[9]: music_tag_format1,
                    csv_title[10]: str(Name_format),
                    csv_title[11]: music_tag_chorus,
                    csv_title[12]: music_file_path,
                    }

        # "候选区域":music_list[num1][music_candidate],
        # print(csv_dict_i)
        csv_dict.append(csv_dict_i)


def music_tag_detection():
    try:
        print("\033[1;35m运行识别标签music_tag_detection\033[0m")
        global num1, lyric1, Name_format, f, music_tag_format1, music_tag_album, music_tag_year, music_tag_chorus, music_tag_tracknumber, music_tag_albumartist
        f=music_tag.load_file("{}".format(music_list_AP[num1]))
        music_info=str(f['artist'])
        music_info=music_info.replace("/", ";")
        music_info=music_info.replace("&", ";")
        music_info=music_info.replace("|", ";")
        m=[]  # 标签中歌手分离，判断是否合唱
        m.append(music_info.split(";"))
        print("\033[33m标签[歌手s]:\033[0m" + str(m))
        # time.sleep(1)
        music_tag_format1="文件名或标签有错误"
        music_tag_album="[none]"
        music_tag_year="[none]"
        music_tag_chorus="合唱"
        music_tag_tracknumber="[none]"
        music_tag_albumartist="[none]"
        for i in range(len(m[0])):
            m[0][i]=m[0][i] + " - " + str(f['title'])
            print("\033[33m标签[歌手 - 歌曲]：\033[0m" + str(m[0][i]) + " \033[33m文件名：\033[0m" + str(music_list[num1][-1]))
            if i == 0:
                if m[0][i] in music_list[num1][-1]:
                    if str(f['albumartist']) + " - " + str(f['title']) in music_list[num1][-1]:
                        music_tag_format1="正确"
            elif m[0][i] in music_list[num1][-1]:
                if str(f['albumartist']) + " - " + str(f['title']) in music_list[num1][-1]:
                    music_tag_format1="正确"

        if len(m[0]) == 1:
            music_tag_chorus="个人"
        if music_tag_chorus == "合唱":
            if ";" in music_list[num1][-1]:
                music_tag_chorus="合唱[文件名歌手只能保留一个]"
            if "&" in music_list[num1][-1]:
                music_tag_chorus="合唱[文件名歌手只能保留一个]"
        if str(f['album']) != "":
            try:
                music_tag_album=str(f['album'])
                Name_format="专辑目录名错误"
                music_artistq=[]
                music_artistq.append(music_list[num1][music_artist].split(" - "))
                print("\033[33m标签[专辑]：\033[0m" + str(f['album']) + "\033[33m专辑文件夹名称：\033[0m" + str(music_artistq))
                if music_artistq[0][0] + " - " in music_list[num1][-1]:
                    if str(f['album']) in music_list[num1][-2]:
                        Name_format="正确"
            except Exception as err:
                print(err)
                print("album标签出错")
        if str(f['albumartist']) != "":
            try:
                music_tag_albumartist=str(f['albumartist'])
            except Exception as err:
                print(err)
                print("albumartist标签出错")
        if str(f['year']) != "":
            try:
                music_tag_year=str(f['year'])
            except Exception as err:
                print(err)
                print("year标签出错")
        if str(f['tracknumber']) != "":
            nt=""
            try:
                if int(f['tracknumber']) in (0, 1, 2, 3, 4, 5, 6, 7, 8, 9):
                    nt="0"
                music_tag_tracknumber=nt + str(f['tracknumber'])
            except Exception as err:
                print(err)
                print("轨道号标签出错")

    except Exception as err:
        print(err)
        print("\033[32;0m错误\033[0m")

    finally:
        zxcv=0


def creatPage_ALL():
    print("\033[1;32m----------------------\033[0m")
    print("\033[1;32m运行creatPage_ALL识别\033[0m")
    global num1, lyric1, Name_format, f, music_tag_format1, music_tag_album, music_tag_year
    # for num1 in range(len(music_list) - 1):
    # print(music_list[num1])
    print(music_list_AP[num1])
    bmm=0
    lyric1="无"

    if "flac" in music_list[num1][-1]:
        for i in range(len(music_list)):
            t=music_list[i][-1].replace(".lrc", ".ok")
            if music_list[num1][-1].replace(".flac", ".ok") == t:
                lyric1="有"
        music_tag_detection()
        try:
            createPage(databaseId, headers)
        except Exception as err:
            print(err)
            print("重试")
            bmm+=1
            time.sleep(1)
            if bmm < 3:
                creatPage_ALL()

    elif "mp3" in music_list[num1][-1]:
        for i in range(len(music_list)):
            t=music_list[i][-1].replace(".lrc", ".ok")
            if music_list[num1][-1].replace(".mp3", ".ok") == t:
                lyric1="有"
        music_tag_detection()
        try:
            createPage(databaseId, headers)
        except Exception as err:
            print(err)
            print("重试")
            time.sleep(1)
            creatPage_ALL()
    elif ".lrc" in music_list[num1][-1]:
        encode_info=get_encode_info(music_list_AP[num1])  # 获取文件编码
        print(encode_info)
        encodeFile2Utf8(music_list_AP[num1])  # 转文件编码到utf8格式
        print("\033[33m[lrc歌词文件]\033[0m修复乱码，utf-8")
    else:
        print("\033[1;33m[非flac、mp3文件跳过]\033[0m" + music_list[num1][-1])


find_music()
print(list(music_list[0]))
print("路径深度:" + str(len(music_list[0])))
music_artist=len(list(music_list[0])) - 2
music_area=music_artist - 1
music_candidate=music_area - 1

num2=len(music_list)
num1=0
while num1 < num2:
    try:
        creatPage_ALL()
    except Exception as err:
        num1=num1 - 1
        num_true=num_true - 1
        time.sleep(5)
    num1=num1 + 1

# 制作表格
print("\033[1;34m--------------------\033[0m")
print("\033[1;34m制作csv\033[0m")
employee_info=csv_title
# print(csv_dict)
new_dict=csv_dict
with open(csv_file, 'w', encoding='utf-8') as csvfile:
    writer=csv.DictWriter(csvfile, fieldnames=employee_info)
    writer.writeheader()
    writer.writerows(new_dict)

# csv重编码，修复乱码
path=csv_file
# read input file
with codecs.open(path, 'r', encoding='utf-8') as file:
    lines=file.read()

# write output file
with codecs.open(path, 'w', encoding='utf_8_sig') as file:
    file.write(lines)


# csv 转 xlsx
def xlsx_to_csv_pd():
    data_xls=pd.read_csv(path, index_col=0)
    data_xls.to_excel(xlsx_file, encoding='utf-8')


xlsx_to_csv_pd()

# 找“专辑目录名错误”的excel位置
print("\033[1;34m渲染excel\033[0m")
file_name=xlsx_file
ex1=openpyxl.load_workbook(file_name)
sheet=ex1.active
excel_tag=[]
excel_tag1=[]
excel_tag2=[]
excel_tag3=[]
excel_tag4=[]
excel_tag5=[]
excel_tag6=[]
for row in sheet.iter_rows(min_row=1, max_row=num_true + 1,
                           min_col=1, max_col=14):
    for cell in row:
        if cell.value == "专辑目录名错误":
            # print(cell.coordinate)
            (row, col)=xlsxwriter.workbook.xl_cell_to_rowcol(cell.coordinate)
            # print(str(row)+","+str(col))
            excel_tag.append([int(row) + 1, int(col) + 1, cell.coordinate])
        if cell.value == "文件名或标签有错误":
            # print(cell.coordinate)
            (row1, col1)=xlsxwriter.workbook.xl_cell_to_rowcol(cell.coordinate)
            # print(str(row)+","+str(col))
            excel_tag1.append([int(row1) + 1, int(col1) + 1, cell.coordinate])
        if cell.value == "无":
            # print(cell.coordinate)
            (row2, col2)=xlsxwriter.workbook.xl_cell_to_rowcol(cell.coordinate)
            # print(str(row)+","+str(col))
            excel_tag2.append([int(row2) + 1, int(col2) + 1, cell.coordinate])
        if cell.value == "合唱":
            # print(cell.coordinate)
            (row3, col3)=xlsxwriter.workbook.xl_cell_to_rowcol(cell.coordinate)
            # print(str(row)+","+str(col))
            excel_tag3.append([int(row3) + 1, int(col3) + 1, cell.coordinate])
        if cell.value == "合唱[文件名歌手只能保留一个]":
            # print(cell.coordinate)
            (row4, col4)=xlsxwriter.workbook.xl_cell_to_rowcol(cell.coordinate)
            # print(str(row)+","+str(col))
            excel_tag4.append([int(row4) + 1, int(col4) + 1, cell.coordinate])
        if cell.value == "[none]":
            # print(cell.coordinate)
            (row5, col5)=xlsxwriter.workbook.xl_cell_to_rowcol(cell.coordinate)
            # print(str(row)+","+str(col))
            excel_tag5.append([int(row5) + 1, int(col5) + 1, cell.coordinate])

for row in sheet.iter_rows(min_row=1, max_row=num_true + 1,
                           min_col=2, max_col=2):  # 地区
    for cell in row:
        if cell.value == "华语":
            # print(cell.coordinate)
            (row6, col6)=xlsxwriter.workbook.xl_cell_to_rowcol(cell.coordinate)
            # print(str(row)+","+str(col))
            excel_tag6.append([int(row6) + 1, int(col6) + 1, cell.coordinate])

# print(excel_tag)
# [[4, 10, 'J4'], [15, 10, 'J15'], [16, 10, 'J16'], [17, 10, 'J17'], [26, 10, 'J26'], [27, 10, 'J27'], [28, 10, 'J28'], [63, 10, 'J63'], [64, 10, 'J64'], [65, 10, 'J65'], [66, 10, 'J66'], [67, 10, 'J67'], [68, 10, 'J68']]
print(len(excel_tag))
title="Sheet1"


def Textcolor(file_name, title):
    global num_true
    wk=openpyxl.load_workbook(file_name)  # 加载已经存在的excel
    sheet1=wk[title]  # wk[wk_name[0]]#title名称
    ws=wk.active
    ws.auto_filter.ref="A1:L1"  # 开启筛选按钮
    if len(excel_tag) != 0:
        print("\033[31m标签专辑目录名错误\033[0m")
        print(excel_tag)
        for i in range(len(excel_tag)):
            Color=['ffc7ce', '9c0006']  # 【背景，字体】红
            fille=PatternFill('solid', fgColor=Color[0])  # 设置填充颜色为 橙色
            font=Font(u'宋体', size=11, bold=False, italic=False, strike=False, color=Color[1])  # 设置字体样式
            sheet1.cell(row=excel_tag[i][0], column=excel_tag[i][1], value="").fill=fille  # 序列
            sheet1.cell(row=excel_tag[i][0], column=excel_tag[i][1], value="专辑目录名错误").font=font  # 序列
    if len(excel_tag1) != 0:
        print("\033[31m文件名或标签有错误\033[0m")
        print(excel_tag1)
        for i in range(len(excel_tag1)):
            Color=['ffc7ce', '9c0006']  # 【背景，字体】红
            fille=PatternFill('solid', fgColor=Color[0])  # 设置填充颜色为 橙色
            font=Font(u'宋体', size=11, bold=False, italic=False, strike=False, color=Color[1])  # 设置字体样式
            sheet1.cell(row=excel_tag1[i][0], column=excel_tag1[i][1], value="").fill=fille  # 序列
            sheet1.cell(row=excel_tag1[i][0], column=excel_tag1[i][1], value="文件名或标签有错误").font=font  # 序列
    if len(excel_tag2) != 0:
        print("\033[33m部分歌曲无歌词\033[0m")
        print(excel_tag2)
        for i in range(len(excel_tag2)):
            Color=['F5DFCB', '9c0006']  # 【背景，字体】红
            fille=PatternFill('solid', fgColor=Color[0])  # 设置填充颜色为 橙色
            font=Font(u'宋体', size=11, bold=False, italic=False, strike=False, color=Color[1])  # 设置字体样式
            sheet1.cell(row=excel_tag2[i][0], column=excel_tag2[i][1], value="").fill=fille  # 序列
            sheet1.cell(row=excel_tag2[i][0], column=excel_tag2[i][1], value="无").font=font  # 序列
    if len(excel_tag3) != 0:
        print(3)
        print(excel_tag3)
        for i in range(len(excel_tag3)):
            Color=['dfecdc', '006122']  # 【背景，字体】浅绿
            fille=PatternFill('solid', fgColor=Color[0])  # 设置填充颜色为 橙色
            font=Font(u'宋体', size=11, bold=False, italic=False, strike=False, color=Color[1])  # 设置字体样式
            sheet1.cell(row=excel_tag3[i][0], column=excel_tag3[i][1], value="").fill=fille  # 序列
            sheet1.cell(row=excel_tag3[i][0], column=excel_tag3[i][1], value="合唱").font=font  # 序列
    if len(excel_tag4) != 0:
        print(4)
        print(excel_tag4)
        for i in range(len(excel_tag4)):
            Color=['ffeb9c', '9c571a']  # 【背景，字体】黄
            fille=PatternFill('solid', fgColor=Color[0])  # 设置填充颜色为 橙色
            font=Font(u'宋体', size=11, bold=False, italic=False, strike=False, color=Color[1])  # 设置字体样式
            sheet1.cell(row=excel_tag4[i][0], column=excel_tag4[i][1], value="").fill=fille  # 序列
            sheet1.cell(row=excel_tag4[i][0], column=excel_tag4[i][1], value="合唱[文件名歌手只能保留一个]").font=font  # 序列
    if len(excel_tag5) != 0:
        print("\033[31m有标签信息未填写\033[0m")
        print(excel_tag5)
        for i in range(len(excel_tag5)):
            Color=['ffeb9c', '9c571a']  # 【背景，字体】黄
            fille=PatternFill('solid', fgColor=Color[0])  # 设置填充颜色为 橙色
            font=Font(u'宋体', size=11, bold=False, italic=False, strike=False, color=Color[1])  # 设置字体样式
            sheet1.cell(row=excel_tag5[i][0], column=excel_tag5[i][1], value="").fill=fille  # 序列
            sheet1.cell(row=excel_tag5[i][0], column=excel_tag5[i][1], value="[none]").font=font  # 序列

    wk.save(file_name)  # 保存excel


Textcolor(file_name, title)

wb=openpyxl.load_workbook(xlsx_file)
ws=wb[wb.sheetnames[0]]  # 打开第一个sheet
ws.column_dimensions['B'].width=10
ws.column_dimensions['C'].width=14
ws.column_dimensions['D'].width=30
ws.column_dimensions['E'].width=6.5
ws.column_dimensions['F'].width=30
ws.column_dimensions['G'].width=14
ws.column_dimensions['H'].width=6.5
ws.column_dimensions['I'].width=10
ws.column_dimensions['J'].width=10
ws.column_dimensions['K'].width=14
ws.column_dimensions['L'].width=10
ws.column_dimensions['M'].width=122
ws.column_dimensions['N'].width=23
wb.save(xlsx_file)

print("------------------------------")
os.remove(music_list_cache1)
os.remove(music_list_cache2)
print("\033[1;34m审核表已生成！！\033[0;33m请查看该专辑文件夹里的\033[1;32m【自动审核表.xlsx】\033[0m")
print(" ")
print(xlsx_file)
print(" ")
input("-----按enter，即可关闭窗口-----")
