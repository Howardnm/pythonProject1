#!/usr/bin/env python
# -*- coding: utf-8 -*-
import music_tag
import os
import requests
import json
import time
import re
import csv
import zlib
import codecs
import pandas as pd
import openpyxl
import xlsxwriter
from openpyxl.styles import Font  # 导入字体模块
from openpyxl.styles import PatternFill  # 导入填充模块

# pip3 install requests


token='secret_DkYKff6IQJdGJTqgFslKQqGpLtsu44RiH3Vsa2nhYjO'

databaseId='7aaf52b505fd4b86ac164ad52d0f8570'

headers={
    "Authorization": "Bearer " + token,
    "Content-Type": "application/json",
    "Notion-Version": "2022-06-28",  # 2021-05-13
    'Cache-Control': 'no-cache'
}

music_files='/volume1/Resilio Sync/folders/plex_music候选区'
music_list_cache1="/volume1/Resilio Sync/folders/foo3.txt"
music_list_cache2="/volume1/Resilio Sync/folders/foo4.txt"
music_list_cache3="/volume1/Resilio Sync/folders/foo4.1.txt"
music_artist=7  # file:6 file1:7
today=time.strftime("%Y/%m/%d %H:%M", time.localtime())
print(today)
csv_title=["ID", "候选区域", "地区", "歌手", "歌曲", "专辑", "年份", "歌词", "标签", "命名格式", "合 唱", "文件路径", today]
csv_file='/volume1/Resilio Sync/folders/plex_music候选区/auto_check.csv'
xlsx_file='/volume1/Resilio Sync/folders/plex_music候选区/自动审核表[30分钟刷新].xlsx'

# ----------------------------------------------------------------
music_area=music_artist - 1
music_candidate=music_area - 1

music_list=[]  # 读取绝对路径，识别"/"，将列表内元素细分，形成套娃列表
music_list_AP=[]  # 读取绝对路径，写入列表
csv_dict=[]
num=0
GB_size=""
json_file1=""
music_tag_format1, music_tag_album, music_tag_year="", "", ""


# 遍历本地音乐文件

def find_music():
    print("运行find_music")
    list1=[]

    def gci(filepath):
        ppp=''
        # 遍历filepath下所有文件，包括子目录
        files=os.listdir(filepath)
        for fi in files:
            fi_d=os.path.join(filepath, fi)
            if os.path.isdir(fi_d):
                gci(fi_d)
            else:
                # print(os.path.join(filepath, fi_d))
                ppp=os.path.join(filepath, fi_d)
                list1.append(ppp)
                # print(list1)

    # 递归遍历/root目录下所有文件
    gci(music_files)

    # 遍历后的绝对路径，写入txt
    fo=open(music_list_cache1, "w+", encoding='utf8')
    for line in list1:
        fo.writelines(line + "\n")
    fo.close()
    # print(list1)

    # 删除包含".sync .txt .jpg .csv"的行
    re_str=['.sync', '.txt', '.jpg', '.csv', '.xlsx', '@eaDir']
    for i in range(len(re_str)):
        list2=[]
        matchPattern=re.compile(r'{}'.format(re_str[i]))
        file=open(music_list_cache1, 'r')
        while 1:
            line=file.readline()
            if not line:
                print("读取文件结束")
                break
            elif matchPattern.search(line):
                pass
            else:
                list2.append(line)
        file.close()
        file=open(music_list_cache1, 'w')
        for i in list2:
            file.write(i)
        file.close()

    file=open(music_list_cache1, 'w')
    for i in list2:
        file.write(i)
    file.close()

    # 读取绝对路径，写入列表
    fo=open(music_list_cache1, "r")
    for line in fo:
        line=line.replace("\n", "")
        music_list_AP.append(line)
    fo.close()
    print("----------------")
    print(music_list_AP)
    # 读取绝对路径，识别"/"，将列表内元素细分，形成套娃列表
    fo=open(music_list_cache1, "r")
    for line in fo:
        line=line.replace("\n", "")
        music_list.append(line.split("/"))
    print(music_list)
    fo.close()
    # 整理后的列表重新写入
    fo=open(music_list_cache2, "w+", encoding='utf8')
    fo.write(str(music_list))
    fo.close()


# 检索文件夹大小
def osos(folder_path):
    full_size=sum(sum(os.path.getsize(os.path.join(parent, file)) for file in files) for parent, dirs, files in
                  os.walk(folder_path))
    global GB_size
    GB_size="%.2f GB" % (full_size / 1024 / 1024 / 1024)
    print("%.2f GB" % (full_size / 1024 / 1024 / 1024))


# notion

# 下载notion-json
def readDatabase(databaseId, headers):
    print("执行readDatabase")
    global json_file1
    readUrl=f"https://api.notion.com/v1/databases/{databaseId}/query"

    res=requests.request("POST", readUrl, headers=headers)
    data=res.json()
    print(res.status_code)
    print(res.text)
    json_file1=res.text
    # print(json_file1)
    with open('./db.json', 'w', encoding='utf8') as f:
        json.dump(data, f, ensure_ascii=False)


# 下载notion-json，提取每一行的id
def removeLine():
    print("执行removeLine")
    global num, num4, token, databaseId, headers, json_file1, dict1
    try:
        test1=dict1.get('results')[0]
    except IndexError:
        print("已清空列表")
    else:
        try:
            data={'archived': True}
            page_id=dict1.get('results')[num].get('id')  # "提取每一行id"
            print(dict1.get('results')[num].get('id'))
            response=requests.patch(
                "https://api.notion.com/v1/pages/{}".format(page_id),
                json=data,
                headers={"Authorization": "Bearer " + token, "Notion-Version": "2022-06-28"},
            )
        except IndexError:
            print("结束")
            time.sleep(5)
            removeLine()
        else:
            num=num + 1
            if num < num4:
                print("num=" + str(num))
                removeLine()
            if num == num4:
                try:
                    readDatabase(databaseId, headers)  # 下载notion-json
                    dict1=json.loads(json_file1)  # json to dict
                    print("转换成字典")
                    print(dict1)
                    num4=len(dict1.get('results'))
                    test1=dict1.get('results')[0]
                except IndexError:
                    print("已清空列表【最后确认】")
                else:
                    num=0
                    print("num4=" + str(num4))
                    removeLine()


def removeLine_All():
    print("执行removeLine_All")
    global num, num4, token, databaseId, headers, json_file1, dict1
    readDatabase(databaseId, headers)  # 下载notion-json
    dict1=json.loads(json_file1)  # json to dict
    print("转换成字典")
    print(dict1)  # 打印dict
    num4=len(dict1.get('results'))
    removeLine()  # 删除每一行


def removePage(pageId, headers, databaseId, token):
    updateUrl=f"https://api.notion.com/v1/databases/{databaseId}"
    headers={
        "Accept": "application/json",
        "Notion-Version": "2022-06-28",
        "Content-Type": "application/json",
        "Authorization": "Bearer " + token,
    }
    updateData={
        "properties": {
            "地区": {
                "multi_select": {
                    "options": [
                        {
                            "name": 'null'
                        },
                    ]
                }
            },
            "歌手": {
                "multi_select": {
                    "options": [
                        {
                            "name": 'null'
                        },
                    ]
                }
            },
            "歌词": {
                "multi_select": {
                    "options": [
                        {
                            "name": 'null'
                        },
                    ]
                }
            }

        }
    }

    data=json.dumps(updateData)
    print(str(data))

    response=requests.request("PATCH", updateUrl, headers=headers, data=data)
    print(response.status_code)
    print(response.text)


num_true=0


def createPage(databaseId, headers):
    print("运行createPage")
    global num_true, num1, lyric1, Name_format, music_tag_format1, music_tag_album, music_tag_year, music_tag_chorus
    print(num1)
    num_true+=1
    print(num_true)
    music_file_path=f"/{music_list[num1][-4]}/{music_list[num1][-3]}/{music_list[num1][-2]}/{music_list[num1][-1]}"
    # music_CRC32=str(zlib.crc32(music_file_path.encode('utf8')))
    # print("music_CRC32=" + music_CRC32)

    music_CRC32=num_true  # ID换回序号
    createUrl='https://api.notion.com/v1/pages'

    newPageData={
        "parent": {"database_id": databaseId},
        "properties": {
            "ID": {
                "title": [
                    {
                        "text": {
                            "content": music_CRC32,
                        },
                    }
                ]
            },
            "候选区域": {
                "multi_select": [
                    {
                        "name": music_list[num1][music_candidate]
                    }
                ]
            },
            "地区": {
                "multi_select": [
                    {
                        "name": music_list[num1][music_area]
                    }
                ]
            },
            "歌手": {
                "multi_select": [
                    {
                        "name": str(f['artist'])
                    }
                ]
            },
            "歌曲": {
                "rich_text": [
                    {
                        "text": {
                            "content": str(f['title'])
                        }
                    }
                ]
            },
            "歌词": {
                "multi_select": [
                    {
                        "name": str(lyric1)
                    }
                ]
            },
            "命名格式": {
                "multi_select": [
                    {
                        "name": str(Name_format)
                    }
                ]
            },
            "专辑": {
                "multi_select": [
                    {
                        "name": music_tag_album
                    }
                ]
            },
            "年份": {
                "multi_select": [
                    {
                        "name": "{}.".format(music_tag_year)
                    }
                ]
            },
            "标签": {
                "multi_select": [
                    {
                        "name": music_tag_format1
                    }
                ]
            },
            "文件路径": {
                "rich_text": [
                    {
                        "text": {
                            "content": music_file_path,
                        }
                    }
                ]
            },
            "守护者": {
                "multi_select": [
                    {
                        "name": "howard"
                    }
                ]
            }
        }
    }
    to_csv=1
    if to_csv == 1:
        print("一首歌的信息导入字典")
        global csv_dict
        global csv_title
        csv_dict_i={csv_title[0]: music_CRC32,
                    csv_title[1]: music_list[num1][music_candidate],
                    csv_title[2]: music_list[num1][music_area],
                    csv_title[3]: str(f['artist']),
                    csv_title[4]: music_tag_t_num + str(f['title']),
                    csv_title[5]: music_tag_album,
                    csv_title[6]: "{}.".format(music_tag_year),
                    csv_title[7]: str(lyric1),
                    csv_title[8]: music_tag_format1,
                    csv_title[9]: str(Name_format),
                    csv_title[10]: music_tag_chorus,
                    csv_title[11]: music_file_path,
                    }

        # "候选区域":music_list[num1][music_candidate],
        print(csv_dict_i)
        csv_dict.append(csv_dict_i)

    post_creat=0
    if post_creat == 1:
        data=json.dumps(newPageData)
        print(str(data))
        res=requests.request("POST", createUrl, headers=headers, data=data)
        print(res.status_code)
        # print(res.text)


def updatePage(pageId, headers, GB_size):
    updateUrl=f"https://api.notion.com/v1/pages/{pageId}"

    updateData={
        "properties": {
            "ID": {
                "rich_text": [
                    {
                        "text": {
                            "content": GB_size
                        }
                    }
                ]
            }
        }
    }

    data=json.dumps(updateData)
    response=requests.request("PATCH", updateUrl, headers=headers, data=data)
    print(response.status_code)
    # print(response.text)


def music_tag_detection():
    try:
        print("运行识别标签music_tag_detection")
        global num1, lyric1, Name_format, f, music_tag_format1, music_tag_album, music_tag_year, music_tag_chorus, music_tag_t_num
        f=music_tag.load_file("{}".format(music_list_AP[num1]))
        music_info=str(f['artist'])
        music_info=music_info.replace("/", ";")
        music_info=music_info.replace("&", ";")
        music_info=music_info.replace("|", ";")
        m=[]
        m.append(music_info.split(";"))
        print(m)
        # time.sleep(1)
        music_tag_format1="文件名未规范化(标签多歌手用/隔开)"
        music_tag_album="缺失"
        music_tag_year="缺失"
        music_tag_chorus="合唱"
        music_tag_t_num="[none]"
        for i in range(len(m[0])):
            m[0][i]=m[0][i] + " - " + str(f['title'])
            if i == 0:
                if m[0][i] in music_list[num1][-1]:
                    music_tag_format1="正确"
            elif m[0][i] == music_list[num1][-1]:
                music_tag_format1="正确"
        if len(m[0]) == 1:
            music_tag_chorus="个人"
        if music_tag_chorus == "合唱":
            if ";" in music_list[num1][-1]:
                music_tag_chorus="合唱[文件名歌手只能保留一个]"
            if "&" in music_list[num1][-1]:
                music_tag_chorus="合唱[文件名歌手只能保留一个]"
        if str(f['album']) != "":
            try:
                music_tag_album=str(f['album'])
            except Exception:
                print("album标签出错")
        if str(f['year']) != "":
            try:
                music_tag_year=str(f['year'])
            except Exception:
                print("year标签出错")

    except Exception as err:
        print(err)
    finally:
        print("-----")


def creatPage_ALL():
    print("运行creatPage_ALL")
    global num1, lyric1, Name_format, f, music_tag_format1, music_tag_album, music_tag_year
    # for num1 in range(len(music_list) - 1):
    print(music_list[num1])
    print(music_list_AP[num1])
    lyric1="无"
    Name_format="不规范"
    if music_list[num1][music_artist] + " - " in music_list[num1][-1]:
        Name_format="正确"
    if "flac" in music_list[num1][-1]:
        for i in range(len(music_list) - 1):
            t=music_list[i][-1].replace(".lrc", ".ok")
            if music_list[num1][-1].replace(".flac", ".ok") == t:
                lyric1="有"
        music_tag_detection()
        try:
            createPage(databaseId, headers)
        except Exception:
            print("重试")
            time.sleep(1)
            creatPage_ALL()

    elif "mp3" in music_list[num1][-1]:
        for i in range(len(music_list) - 1):
            t=music_list[i][-1].replace(".lrc", ".ok")
            if music_list[num1][-1].replace(".mp3", ".ok") == t:
                lyric1="有"
        music_tag_detection()
        try:
            createPage(databaseId, headers)
        except Exception:
            print("重试")
            time.sleep(1)
            creatPage_ALL()
    else:
        print("[非flac、mp3文件跳过]" + music_list[num1][-1])


# readDatabase(databaseId, headers)

# plex_music
# pageId='832b6bfc38e946af9ea15732dfdfdf29'
# updatePage(pageId, headers, GB_size)

# removePage(pageId, headers, databaseId, token)

# removeLine_All()

find_music()
print(len(music_list))

num2=len(music_list) - 1
num1=0
while num1 < num2:
    try:
        creatPage_ALL()
    except Exception:
        num1=num1 - 1
        num_true=num_true - 1
        time.sleep(5)
    num1=num1 + 1

# 制作表格

print("制作csv")
employee_info=csv_title
print(csv_dict)
new_dict=csv_dict
with open(csv_file, 'w') as csvfile:
    writer=csv.DictWriter(csvfile, fieldnames=employee_info)
    writer.writeheader()
    writer.writerows(new_dict)

# csv重编码，修复乱码
path=csv_file
# read input file
with codecs.open(path, 'r', encoding='utf-8') as file:
    lines=file.read()

# write output file
with codecs.open(path, 'w', encoding='utf_8_sig') as file:
    file.write(lines)


# csv 转 xlsx
def xlsx_to_csv_pd():
    data_xls=pd.read_csv(path, index_col=0)
    data_xls.to_excel(xlsx_file, encoding='utf-8')


xlsx_to_csv_pd()

# 找“不规范”的excel位置
file_name=xlsx_file
ex1=openpyxl.load_workbook(file_name)
sheet=ex1.active
excel_tag=[]
excel_tag1=[]
excel_tag2=[]
excel_tag3=[]
excel_tag4=[]
excel_tag5=[]
excel_tag6=[]
for row in sheet.iter_rows(min_row=1, max_row=num_true + 1,
                           min_col=1, max_col=11):
    for cell in row:
        if cell.value == "不规范":
            # print(cell.coordinate)
            (row, col)=xlsxwriter.workbook.xl_cell_to_rowcol(cell.coordinate)
            # print(str(row)+","+str(col))
            excel_tag.append([int(row) + 1, int(col) + 1, cell.coordinate])
        if cell.value == "文件名未规范化(标签多歌手用/隔开)":
            # print(cell.coordinate)
            (row1, col1)=xlsxwriter.workbook.xl_cell_to_rowcol(cell.coordinate)
            # print(str(row)+","+str(col))
            excel_tag1.append([int(row1) + 1, int(col1) + 1, cell.coordinate])
        if cell.value == "无":
            # print(cell.coordinate)
            (row2, col2)=xlsxwriter.workbook.xl_cell_to_rowcol(cell.coordinate)
            # print(str(row)+","+str(col))
            excel_tag2.append([int(row2) + 1, int(col2) + 1, cell.coordinate])
        if cell.value == "合唱":
            # print(cell.coordinate)
            (row3, col3)=xlsxwriter.workbook.xl_cell_to_rowcol(cell.coordinate)
            # print(str(row)+","+str(col))
            excel_tag3.append([int(row3) + 1, int(col3) + 1, cell.coordinate])
        if cell.value == "合唱[文件名歌手只能保留一个]":
            # print(cell.coordinate)
            (row4, col4)=xlsxwriter.workbook.xl_cell_to_rowcol(cell.coordinate)
            # print(str(row)+","+str(col))
            excel_tag4.append([int(row4) + 1, int(col4) + 1, cell.coordinate])

for row in sheet.iter_rows(min_row=1, max_row=num_true + 1,
                           min_col=2, max_col=2):  # 地区
    for cell in row:
        if cell.value == "华语":
            # print(cell.coordinate)
            (row6, col6)=xlsxwriter.workbook.xl_cell_to_rowcol(cell.coordinate)
            # print(str(row)+","+str(col))
            excel_tag6.append([int(row6) + 1, int(col6) + 1, cell.coordinate])

# print(excel_tag)
# [[4, 10, 'J4'], [15, 10, 'J15'], [16, 10, 'J16'], [17, 10, 'J17'], [26, 10, 'J26'], [27, 10, 'J27'], [28, 10, 'J28'], [63, 10, 'J63'], [64, 10, 'J64'], [65, 10, 'J65'], [66, 10, 'J66'], [67, 10, 'J67'], [68, 10, 'J68']]
print(len(excel_tag))
title="Sheet1"


def Textcolor(file_name, title):
    global num_true
    wk=openpyxl.load_workbook(file_name)  # 加载已经存在的excel
    sheet1=wk[title]  # wk[wk_name[0]]#title名称
    ws=wk.active
    ws.auto_filter.ref="A1:L1"  # 开启筛选按钮
    if len(excel_tag) != 0:
        print(0)
        print(excel_tag)
        for i in range(len(excel_tag)):
            Color=['ffc7ce', '9c0006']  # 【背景，字体】红
            fille=PatternFill('solid', fgColor=Color[0])  # 设置填充颜色为 橙色
            font=Font(u'宋体', size=11, bold=False, italic=False, strike=False, color=Color[1])  # 设置字体样式
            sheet1.cell(row=excel_tag[i][0], column=excel_tag[i][1], value="").fill=fille  # 序列
            sheet1.cell(row=excel_tag[i][0], column=excel_tag[i][1], value="不规范").font=font  # 序列
    if len(excel_tag1) != 0:
        print(1)
        print(excel_tag1)
        for i in range(len(excel_tag1)):
            Color=['ffc7ce', '9c0006']  # 【背景，字体】红
            fille=PatternFill('solid', fgColor=Color[0])  # 设置填充颜色为 橙色
            font=Font(u'宋体', size=11, bold=False, italic=False, strike=False, color=Color[1])  # 设置字体样式
            sheet1.cell(row=excel_tag1[i][0], column=excel_tag1[i][1], value="").fill=fille  # 序列
            sheet1.cell(row=excel_tag1[i][0], column=excel_tag1[i][1], value="文件名未规范化(标签多歌手用/隔开)").font=font  # 序列
    if len(excel_tag2) != 0:
        print(2)
        print(excel_tag2)
        for i in range(len(excel_tag2)):
            Color=['F5DFCB', '9c0006']  # 【背景，字体】红
            fille=PatternFill('solid', fgColor=Color[0])  # 设置填充颜色为 橙色
            font=Font(u'宋体', size=11, bold=False, italic=False, strike=False, color=Color[1])  # 设置字体样式
            sheet1.cell(row=excel_tag2[i][0], column=excel_tag2[i][1], value="").fill=fille  # 序列
            sheet1.cell(row=excel_tag2[i][0], column=excel_tag2[i][1], value="无").font=font  # 序列
    if len(excel_tag3) != 0:
        print(3)
        print(excel_tag3)
        for i in range(len(excel_tag3)):
            Color=['dfecdc', '006122']  # 【背景，字体】浅绿
            fille=PatternFill('solid', fgColor=Color[0])  # 设置填充颜色为 橙色
            font=Font(u'宋体', size=11, bold=False, italic=False, strike=False, color=Color[1])  # 设置字体样式
            sheet1.cell(row=excel_tag3[i][0], column=excel_tag3[i][1], value="").fill=fille  # 序列
            sheet1.cell(row=excel_tag3[i][0], column=excel_tag3[i][1], value="合唱").font=font  # 序列
    if len(excel_tag4) != 0:
        print(4)
        print(excel_tag4)
        for i in range(len(excel_tag4)):
            Color=['ffeb9c', '9c571a']  # 【背景，字体】黄
            fille=PatternFill('solid', fgColor=Color[0])  # 设置填充颜色为 橙色
            font=Font(u'宋体', size=11, bold=False, italic=False, strike=False, color=Color[1])  # 设置字体样式
            sheet1.cell(row=excel_tag4[i][0], column=excel_tag4[i][1], value="").fill=fille  # 序列
            sheet1.cell(row=excel_tag4[i][0], column=excel_tag4[i][1], value="合唱[文件名歌手只能保留一个]").font=font  # 序列

    wk.save(file_name)  # 保存excel


Textcolor(file_name, title)

wb=openpyxl.load_workbook(xlsx_file)
ws=wb[wb.sheetnames[0]]  # 打开第一个sheet
ws.column_dimensions['B'].width=12
ws.column_dimensions['C'].width=10
ws.column_dimensions['D'].width=14
ws.column_dimensions['E'].width=30
ws.column_dimensions['F'].width=30
ws.column_dimensions['G'].width=10
ws.column_dimensions['H'].width=10
ws.column_dimensions['I'].width=10
ws.column_dimensions['J'].width=10
ws.column_dimensions['K'].width=10
ws.column_dimensions['L'].width=122
ws.column_dimensions['M'].width=23
wb.save(xlsx_file)

print("------------------------------")
print("------------------------------")
print("------------------------------")
print("------------------------------")
print("所有程序执行完成！！")